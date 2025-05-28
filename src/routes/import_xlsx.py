import os
import uuid
import tempfile
import openpyxl
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.models import db, Employee, WorkDay, WorkPeriod

import_xlsx_bp = Blueprint('import_xlsx', __name__, url_prefix='/import-xlsx')

ALLOWED_EXTENSIONS = {'xlsx'}
UPLOAD_FOLDER = tempfile.gettempdir()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@import_xlsx_bp.route('/')
def index():
    return render_template('import_xlsx/index.html')

@import_xlsx_bp.route('/download-template')
def download_template():
    try:
        # Criar um arquivo Excel de exemplo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo"
        
        # Definir cabeçalhos
        headers = ["Nome Funcionário", "Data", "Hora Extra Inicial", "Hora Extra Final"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Adicionar exemplos
        examples = [
            ["João Silva", "01/03/25", "03:00", "06:00"],
            ["João Silva", "01/03/25", "19:00", "20:00"],
            ["João Silva", "01/03/25", "21:00", "23:30"],
            ["Maria Santos", "02/03/25", "22:00", "23:59"],
            ["Maria Santos", "03/03/25", "00:01", "03:00"]
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')
        
        # Salvar o arquivo temporariamente
        temp_file = os.path.join(UPLOAD_FOLDER, "modelo_importacao_horas_extras.xlsx")
        wb.save(temp_file)
        
        return send_file(temp_file, as_attachment=True, download_name="modelo_importacao_horas_extras.xlsx")
    except Exception as e:
        flash(f'Erro ao gerar modelo: {str(e)}', 'danger')
        return redirect(url_for('import_xlsx.index'))

def validate_record(record_data):
    """Valida um registro de hora extra e retorna o resultado da validação"""
    record = {
        'employee_name': record_data.get('employee_name', '').strip(),
        'date': record_data.get('date', '').strip(),
        'start_time': record_data.get('start_time', '').strip(),
        'end_time': record_data.get('end_time', '').strip(),
        'status': 'valid',
        'message': '',
        'employee_id': None,
        'date_obj': None,
        'start_time_obj': None,
        'end_time_obj': None
    }
    
    # Validar funcionário
    if not record['employee_name']:
        record['status'] = 'error'
        record['message'] = 'Nome do funcionário não pode ser vazio'
        return record
    
    employee = Employee.query.filter_by(name=record['employee_name']).first()
    if not employee:
        record['status'] = 'error'
        record['message'] = f"Funcionário '{record['employee_name']}' não encontrado"
        return record
    
    record['employee_id'] = employee.id
    
    # Validar data
    if not record['date']:
        record['status'] = 'error'
        record['message'] = 'Data não pode ser vazia'
        return record
    
    try:
        # Tentar diferentes formatos de data
        formats = ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d']
        date_parsed = False
        
        for fmt in formats:
            try:
                record['date_obj'] = datetime.strptime(record['date'], fmt).date()
                date_parsed = True
                break
            except ValueError:
                continue
        
        if not date_parsed:
            record['status'] = 'error'
            record['message'] = f"Formato de data inválido: {record['date']}. Use DD/MM/YY"
            return record
    except Exception as e:
        record['status'] = 'error'
        record['message'] = f"Erro ao processar data: {str(e)}"
        return record
    
    # Validar hora inicial
    if not record['start_time']:
        record['status'] = 'error'
        record['message'] = 'Hora Extra Inicial não pode ser vazia'
        return record
    
    try:
        record['start_time_obj'] = datetime.strptime(record['start_time'], '%H:%M').time()
    except Exception as e:
        record['status'] = 'error'
        record['message'] = f"Formato de hora inválido para Hora Extra Inicial: {str(e)}"
        return record
    
    # Validar hora final
    if not record['end_time']:
        record['status'] = 'error'
        record['message'] = 'Hora Extra Final não pode ser vazia'
        return record
    
    try:
        record['end_time_obj'] = datetime.strptime(record['end_time'], '%H:%M').time()
    except Exception as e:
        record['status'] = 'error'
        record['message'] = f"Formato de hora inválido para Hora Extra Final: {str(e)}"
        return record
    
    # Validar intervalo de tempo
    if record['start_time_obj'] > record['end_time_obj']:
        record['status'] = 'warning'
        record['message'] = 'Atenção: Hora final é anterior à hora inicial. Será considerado como período que cruza a meia-noite.'
    
    # Formatar para exibição
    record['date'] = record['date_obj'].strftime('%d/%m/%Y')
    record['start_time'] = record['start_time_obj'].strftime('%H:%M')
    record['end_time'] = record['end_time_obj'].strftime('%H:%M')
    
    return record

def save_record_to_database(record):
    """Salva um registro validado no banco de dados"""
    try:
        # Verificar se já existe um WorkDay para esta data e funcionário
        work_day = WorkDay.query.filter_by(
            employee_id=record['employee_id'],
            date=record['date_obj']
        ).first()
        
        # Se não existir, criar um novo
        if not work_day:
            work_day = WorkDay(
                employee_id=record['employee_id'],
                date=record['date_obj']
            )
            db.session.add(work_day)
            db.session.flush()  # Para obter o ID
        
        # Criar um novo período adicional
        work_period = WorkPeriod(
            work_day_id=work_day.id,
            period_type='additional',
            entry_time=record['start_time_obj'],
            exit_time=record['end_time_obj']
        )
        
        db.session.add(work_period)
        return True
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao salvar registro: {str(e)}', 'danger')
        return False

@import_xlsx_bp.route('/process-manual', methods=['POST'])
def process_manual():
    """Processa um único registro inserido manualmente"""
    try:
        record_data = {
            'employee_name': request.form.get('employee_name', ''),
            'date': request.form.get('date', ''),
            'start_time': request.form.get('start_time', ''),
            'end_time': request.form.get('end_time', '')
        }
        
        record = validate_record(record_data)
        
        if record['status'] == 'error':
            flash(f"Erro: {record['message']}", 'danger')
            return redirect(url_for('import_xlsx.index'))
        
        if save_record_to_database(record):
            db.session.commit()
            flash('Registro de hora extra adicionado com sucesso!', 'success')
            
            import_results = {
                'total_count': 1,
                'success_count': 1,
                'ignored_count': 0
            }
            
            return render_template('import_xlsx/index.html', import_results=import_results)
        else:
            return redirect(url_for('import_xlsx.index'))
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar o registro: {str(e)}', 'danger')
        return redirect(url_for('import_xlsx.index'))

@import_xlsx_bp.route('/process-multiple', methods=['POST'])
def process_multiple():
    """Processa múltiplos registros colados como texto"""
    try:
        multiple_data = request.form.get('multiple_data', '')
        lines = multiple_data.strip().split('\n')
        
        preview_data = []
        has_errors = False
        
        for line in lines:
            if not line.strip():
                continue
                
            parts = [p.strip() for p in line.split(',')]
            
            if len(parts) < 4:
                continue
                
            record_data = {
                'employee_name': parts[0],
                'date': parts[1],
                'start_time': parts[2],
                'end_time': parts[3]
            }
            
            record = validate_record(record_data)
            
            if record['status'] == 'error':
                has_errors = True
                
            preview_data.append(record)
        
        # Gerar ID de sessão único para rastrear estes dados
        session_id = str(uuid.uuid4())
        session[f'import_data_{session_id}'] = preview_data
        
        return render_template('import_xlsx/index.html', 
                              preview_data=preview_data, 
                              session_id=session_id,
                              has_errors=has_errors)
    
    except Exception as e:
        flash(f'Erro ao processar os dados: {str(e)}', 'danger')
        return redirect(url_for('import_xlsx.index'))

@import_xlsx_bp.route('/process', methods=['POST'])
def process_import():
    """Processa os dados previamente validados"""
    try:
        session_id = request.form.get('session_id')
        
        if not session_id or f'import_data_{session_id}' not in session:
            flash('Sessão de importação inválida ou expirada', 'danger')
            return redirect(url_for('import_xlsx.index'))
        
        preview_data = session[f'import_data_{session_id}']
        
        success_count = 0
        ignored_count = 0
        
        for record in preview_data:
            if record['status'] == 'valid' or record['status'] == 'warning':
                if save_record_to_database(record):
                    success_count += 1
                else:
                    ignored_count += 1
            else:
                ignored_count += 1
        
        db.session.commit()
        
        # Limpar dados da sessão
        if f'import_data_{session_id}' in session:
            del session[f'import_data_{session_id}']
        
        import_results = {
            'total_count': len(preview_data),
            'success_count': success_count,
            'ignored_count': ignored_count
        }
        
        flash(f'Importação concluída com sucesso! {success_count} registros importados.', 'success')
        return render_template('import_xlsx/index.html', import_results=import_results)
    
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar a importação: {str(e)}', 'danger')
        return redirect(url_for('import_xlsx.index'))

@import_xlsx_bp.route('/upload', methods=['POST'])
def upload_file():
    """Mantido para compatibilidade, mas não recomendado devido a problemas no ambiente de produção"""
    flash('O upload direto de arquivos está temporariamente desativado. Por favor, use a entrada manual ou múltipla de dados.', 'warning')
    return redirect(url_for('import_xlsx.index'))
