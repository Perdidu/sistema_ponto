from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file
from werkzeug.utils import secure_filename
import os
import uuid
import pandas as pd
from datetime import datetime
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.models import db, Employee, WorkDay, WorkPeriod
from src.utils.import_utils import ImportValidator

import_bp = Blueprint('import', __name__, url_prefix='/import')

ALLOWED_EXTENSIONS = {'xlsx'}
UPLOAD_FOLDER = tempfile.gettempdir()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@import_bp.route('/')
def index():
    return render_template('import/index.html')

@import_bp.route('/download-template')
def download_template():
    # Criar um arquivo Excel de exemplo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"
    
    # Definir cabeçalhos
    headers = ["Nome Funcionário", "Data", "Hora Extra Inicial", "Hora Extra Final"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    # Adicionar exemplos
    examples = [
        ["João Silva", "01/03/25", "03:00", "06:00"],
        ["João Silva", "01/03/25", "19:00", "20:00"],
        ["João Silva", "01/03/25", "21:00", "23:30"],
        ["Maria Santos", "02/03/25", "22:00", "23:59"],
        ["Maria Santos", "03/03/25", "00:01", "03:00"]
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
    
    # Salvar o arquivo temporariamente
    temp_file = os.path.join(UPLOAD_FOLDER, "modelo_importacao_horas_extras.xlsx")
    wb.save(temp_file)
    
    return send_file(temp_file, as_attachment=True, download_name="modelo_importacao_horas_extras.xlsx")

@import_bp.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('Nenhum arquivo selecionado', 'danger')
        return redirect(url_for('import.index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('Nenhum arquivo selecionado', 'danger')
        return redirect(url_for('import.index'))
    
    if not allowed_file(file.filename):
        flash('Formato de arquivo não permitido. Use apenas .xlsx', 'danger')
        return redirect(url_for('import.index'))
    
    # Gerar ID de sessão único para rastrear este upload
    session_id = str(uuid.uuid4())
    
    # Salvar o arquivo temporariamente
    filename = secure_filename(file.filename)
    temp_path = os.path.join(UPLOAD_FOLDER, f"{session_id}_{filename}")
    file.save(temp_path)
    
    try:
        # Ler o arquivo Excel
        df = pd.read_excel(temp_path)
        
        # Verificar se as colunas necessárias existem
        required_columns = ["Nome Funcionário", "Data", "Hora Extra Inicial", "Hora Extra Final"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            flash(f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}", 'danger')
            os.remove(temp_path)
            return redirect(url_for('import.index'))
        
        # Validar e preparar os dados
        validator = ImportValidator()
        preview_data = []
        has_errors = False
        
        for index, row in df.iterrows():
            record_data = {
                'Nome Funcionário': row['Nome Funcionário'],
                'Data': row['Data'],
                'Hora Extra Inicial': row['Hora Extra Inicial'],
                'Hora Extra Final': row['Hora Extra Final']
            }
            
            record = validator.process_record(record_data)
            
            if record['status'] == 'error':
                has_errors = True
                
            preview_data.append(record)
        
        # Armazenar dados na sessão para processamento posterior
        session[f'import_data_{session_id}'] = preview_data
        session[f'import_file_{session_id}'] = temp_path
        
        return render_template('import/index.html', 
                              preview_data=preview_data, 
                              session_id=session_id,
                              has_errors=has_errors)
    
    except Exception as e:
        flash(f'Erro ao processar o arquivo: {str(e)}', 'danger')
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return redirect(url_for('import.index'))

@import_bp.route('/process', methods=['POST'])
def process_import():
    session_id = request.form.get('session_id')
    
    if not session_id or f'import_data_{session_id}' not in session:
        flash('Sessão de importação inválida ou expirada', 'danger')
        return redirect(url_for('import.index'))
    
    preview_data = session[f'import_data_{session_id}']
    temp_path = session[f'import_file_{session_id}']
    
    try:
        validator = ImportValidator()
        result = validator.save_to_database(preview_data)
        
        # Limpar dados da sessão
        if f'import_data_{session_id}' in session:
            del session[f'import_data_{session_id}']
        
        if f'import_file_{session_id}' in session:
            del session[f'import_file_{session_id}']
        
        # Remover arquivo temporário
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        if result['success']:
            flash(result['message'], 'success')
            import_results = {
                'total_count': len(preview_data),
                'success_count': result['success_count'],
                'ignored_count': result['ignored_count']
            }
            return render_template('import/index.html', import_results=import_results)
        else:
            flash(result['message'], 'danger')
            return redirect(url_for('import.index'))
    
    except Exception as e:
        flash(f'Erro ao processar a importação: {str(e)}', 'danger')
        
        # Limpar dados da sessão
        if f'import_data_{session_id}' in session:
            del session[f'import_data_{session_id}']
        
        if f'import_file_{session_id}' in session:
            del session[f'import_file_{session_id}']
        
        # Remover arquivo temporário
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
        return redirect(url_for('import.index'))
